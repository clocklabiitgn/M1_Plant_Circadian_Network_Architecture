#!/usr/bin/env python3
import csv
import os
import argparse

import matplotlib.pyplot as plt
from openpyxl import Workbook


def set_pub_style() -> None:
    plt.rcParams.update({
        "font.family": "Arial",
        "font.size": 13,
        "axes.labelsize": 14,
        "axes.titlesize": 16,
        "legend.fontsize": 13,
        "xtick.labelsize": 11,
        "ytick.labelsize": 11,
        "axes.linewidth": 1,
        "xtick.direction": "out",
        "ytick.direction": "out",
        "xtick.major.size": 4,
        "ytick.major.size": 4,
        "xtick.major.width": 1,
        "ytick.major.width": 1,
        "savefig.bbox": "tight",
    })


def _to_float(value):
    if value is None:
        return None
    s = str(value).strip()
    if s == "" or s.lower() == "nan":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def read_summary_csv(path):
    """Read a pre-computed summary CSV with columns: Parameter, Mean_Abs_Change, Category."""
    data = {}
    with open(path, "r", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            param = (row.get("Parameter") or "").strip()
            if not param:
                continue
            data[param] = {
                "mean_abs_change": _to_float(row.get("Mean_Abs_Change")),
                "category": (row.get("Category") or "").strip(),
            }
    return data


def _assign_combined_class(score):
    if score >= 5.0:
        return "Class I"
    elif score >= 0.5:
        return "Class II"
    else:
        return "Class III"


def plot_dumbbell(ll_data, ld_data, out_path, dpi):
    set_pub_style()

    all_params = sorted(set(ll_data.keys()) | set(ld_data.keys()))
    if not all_params:
        raise SystemExit("No parameters found.")

    # Compute combined score (mean of LL and LD mean_abs_change) for sorting and class
    combined = {}
    for p in all_params:
        vals = []
        if ll_data.get(p, {}).get("mean_abs_change") is not None:
            vals.append(ll_data[p]["mean_abs_change"])
        if ld_data.get(p, {}).get("mean_abs_change") is not None:
            vals.append(ld_data[p]["mean_abs_change"])
        combined[p] = sum(vals) / len(vals) if vals else 0.0

    sorted_params = sorted(all_params, key=lambda p: combined[p], reverse=True)
    x = list(range(len(sorted_params)))
    ll_vals = [ll_data.get(p, {}).get("mean_abs_change") or 0.0 for p in sorted_params]
    ld_vals = [ld_data.get(p, {}).get("mean_abs_change") or 0.0 for p in sorted_params]

    # Class colors for x-tick labels
    category_colors = {"Class I": "#0072B2", "Class II": "#E69F00", "Class III": "#999999"}

    fig_width = min(10.5, max(9.5, len(sorted_params) * 0.34))
    fig, ax = plt.subplots(figsize=(fig_width, 3.7))

    for i in range(len(x)):
        ax.plot([i, i], [ll_vals[i], ld_vals[i]], color="gray", linewidth=1, alpha=0.6, zorder=1)
    ax.scatter(x, ll_vals, color="#0072B2", s=45, label="LL", zorder=3, edgecolors="white", linewidth=0.5)
    ax.scatter(x, ld_vals, color="#E69F00", s=45, label="LD", zorder=3, edgecolors="white", linewidth=0.5)

    ax.set_xticks(x)
    ax.set_xticklabels(sorted_params, rotation=90, ha="center", va="top")
    ax.tick_params(axis="both", pad=4)
    ax.set_xlim(-0.45, len(x) - 0.55)
    ax.set_xlabel("Parameter")
    ax.set_ylabel("Mean |Δ Period| (hours)")
    ax.set_title("Knockout effect of parameters (LL vs LD)", pad=10)
    ax.grid(axis="y", linestyle="--", linewidth=0.5, alpha=0.6)
    ax.axhline(0, color="black", linewidth=0.8, alpha=0.8)
    ax.legend(frameon=False, loc="upper right")

    for tick, p in zip(ax.get_xticklabels(), sorted_params):
        cls = _assign_combined_class(combined[p])
        tick.set_color(category_colors[cls])

    fig.subplots_adjust(left=0.08, right=0.99, top=0.88, bottom=0.52)
    fig.savefig(out_path, dpi=dpi, bbox_inches="tight")
    plt.close(fig)

    return combined


def main():
    parser = argparse.ArgumentParser(description="Plot dumbbell chart of knockout delta means (LL vs LD).")
    parser.add_argument("--ll", default="knockout_analysis_summary_LL.csv", help="LL summary CSV path")
    parser.add_argument("--ld", default="knockout_analysis_summary_LD.csv", help="LD summary CSV path")
    parser.add_argument("--out", default="dumbbell_knockout.png", help="Output PNG path")
    parser.add_argument("--dpi", type=int, default=300)
    args = parser.parse_args()

    ll_data = read_summary_csv(args.ll)
    ld_data = read_summary_csv(args.ld)

    out_dir = os.path.dirname(args.out) or "."
    os.makedirs(out_dir, exist_ok=True)

    combined = plot_dumbbell(ll_data, ld_data, args.out, args.dpi)
    print(f"Saved dumbbell plot: {args.out}")

    # Excel summary
    wb = Workbook()
    ws_ll = wb.active
    ws_ll.title = "LL"
    ws_ld = wb.create_sheet("LD")
    ws_combined = wb.create_sheet("Combined")

    headers = ["Parameter", "Mean_Abs_Change", "Category"]
    for ws in [ws_ll, ws_ld, ws_combined]:
        ws.append(headers)

    for p, v in sorted(ll_data.items()):
        ws_ll.append([p, v["mean_abs_change"], v["category"]])
    for p, v in sorted(ld_data.items()):
        ws_ld.append([p, v["mean_abs_change"], v["category"]])
    for p in sorted(combined):
        ws_combined.append([p, combined[p], _assign_combined_class(combined[p])])

    excel_path = os.path.join(out_dir, "knockout_delta_summary.xlsx")
    wb.save(excel_path)
    print(f"Saved Excel summary: {excel_path}")


if __name__ == "__main__":
    main()
