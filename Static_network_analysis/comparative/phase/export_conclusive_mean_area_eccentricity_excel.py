import csv
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook

# Input files
CSV_LL = Path("mean_metrics_ll.csv")
CSV_LD = Path("mean_metrics_ld.csv")

# Output file
OUT_XLSX = Path("conclusive_mean_area_eccentricity_vs_fold_change.xlsx")

# Column names
PARAM_COL = "parameter"
FOLD_COL = "fold_change"
PARAM_VALUE_COL = "param_value"
BASE_PARAM_VALUE_COL = "base_param_value"
AREA_COL = "mean_area"
ECC_COL = "mean_eccentricity"
DELTA_AREA_COL = "mean_delta_area"
DELTA_ECC_COL = "mean_delta_eccentricity"

REQUIRED_COLUMNS = {
    PARAM_COL,
    FOLD_COL,
    PARAM_VALUE_COL,
    BASE_PARAM_VALUE_COL,
    AREA_COL,
    ECC_COL,
    DELTA_AREA_COL,
    DELTA_ECC_COL,
}


def to_float(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def compute_slope(xs, ys):
    # Ordinary least-squares slope for y = a*x + b.
    pairs = [(x, y) for x, y in zip(xs, ys) if x is not None and y is not None]
    if len(pairs) < 2:
        return None

    x_vals = [p[0] for p in pairs]
    if len(set(x_vals)) < 2:
        return None

    y_vals = [p[1] for p in pairs]
    x_mean = sum(x_vals) / len(x_vals)
    y_mean = sum(y_vals) / len(y_vals)

    num = sum((x - x_mean) * (y - y_mean) for x, y in pairs)
    den = sum((x - x_mean) ** 2 for x in x_vals)
    if den == 0:
        return None
    return num / den


def load_rows(csv_path):
    with open(csv_path, newline="") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None:
            raise ValueError(f"No header found in {csv_path}")
        missing = sorted(REQUIRED_COLUMNS - set(reader.fieldnames))
        if missing:
            raise ValueError(f"Missing columns in {csv_path}: {missing}")
        return list(reader)


def sort_key(row):
    param = row.get(PARAM_COL, "")
    fold = to_float(row.get(FOLD_COL))
    if fold is None:
        fold = float("inf")
    return (param, fold)


def prepare_condition_sheet_rows(rows, condition_label):
    output = []
    for row in sorted(rows, key=sort_key):
        output.append(
            [
                row.get(PARAM_COL),
                to_float(row.get(PARAM_VALUE_COL)),
                to_float(row.get(BASE_PARAM_VALUE_COL)),
                to_float(row.get(FOLD_COL)),
                to_float(row.get(AREA_COL)),
                to_float(row.get(ECC_COL)),
                to_float(row.get(DELTA_AREA_COL)),
                to_float(row.get(DELTA_ECC_COL)),
                condition_label,
            ]
        )
    return output


def make_join_key(row):
    return (row.get(PARAM_COL), to_float(row.get(FOLD_COL)))


def build_conclusive_rows(rows_ll, rows_ld):
    map_ll = {make_join_key(r): r for r in rows_ll}
    map_ld = {make_join_key(r): r for r in rows_ld}
    keys = sorted(set(map_ll) | set(map_ld), key=lambda k: (k[0], float("inf") if k[1] is None else k[1]))

    rows = []
    for key in keys:
        ll = map_ll.get(key, {})
        ld = map_ld.get(key, {})

        mean_area_ll = to_float(ll.get(AREA_COL))
        mean_area_ld = to_float(ld.get(AREA_COL))
        mean_ecc_ll = to_float(ll.get(ECC_COL))
        mean_ecc_ld = to_float(ld.get(ECC_COL))
        delta_area_ll = to_float(ll.get(DELTA_AREA_COL))
        delta_area_ld = to_float(ld.get(DELTA_AREA_COL))
        delta_ecc_ll = to_float(ll.get(DELTA_ECC_COL))
        delta_ecc_ld = to_float(ld.get(DELTA_ECC_COL))

        rows.append(
            [
                key[0],
                key[1],
                to_float(ll.get(PARAM_VALUE_COL)),
                to_float(ld.get(PARAM_VALUE_COL)),
                to_float(ll.get(BASE_PARAM_VALUE_COL)),
                to_float(ld.get(BASE_PARAM_VALUE_COL)),
                mean_area_ll,
                mean_area_ld,
                mean_ecc_ll,
                mean_ecc_ld,
                delta_area_ll,
                delta_area_ld,
                delta_ecc_ll,
                delta_ecc_ld,
                None if (mean_area_ll is None or mean_area_ld is None) else mean_area_ld - mean_area_ll,
                None if (mean_ecc_ll is None or mean_ecc_ld is None) else mean_ecc_ld - mean_ecc_ll,
                None if (delta_area_ll is None or delta_area_ld is None) else delta_area_ld - delta_area_ll,
                None if (delta_ecc_ll is None or delta_ecc_ld is None) else delta_ecc_ld - delta_ecc_ll,
            ]
        )
    return rows


def build_rate_summary_rows(rows_ll, rows_ld):
    by_param_ll = defaultdict(list)
    by_param_ld = defaultdict(list)

    for row in rows_ll:
        by_param_ll[row.get(PARAM_COL, "")].append(row)
    for row in rows_ld:
        by_param_ld[row.get(PARAM_COL, "")].append(row)

    params = sorted(set(by_param_ll) | set(by_param_ld))
    output = []

    for param in params:
        ll_rows = by_param_ll.get(param, [])
        ld_rows = by_param_ld.get(param, [])

        x_ll = [to_float(r.get(FOLD_COL)) for r in ll_rows]
        x_ld = [to_float(r.get(FOLD_COL)) for r in ld_rows]

        slope_delta_area_ll = compute_slope(x_ll, [to_float(r.get(DELTA_AREA_COL)) for r in ll_rows])
        slope_delta_area_ld = compute_slope(x_ld, [to_float(r.get(DELTA_AREA_COL)) for r in ld_rows])
        slope_delta_ecc_ll = compute_slope(x_ll, [to_float(r.get(DELTA_ECC_COL)) for r in ll_rows])
        slope_delta_ecc_ld = compute_slope(x_ld, [to_float(r.get(DELTA_ECC_COL)) for r in ld_rows])

        output.append(
            [
                param,
                slope_delta_area_ll,
                slope_delta_area_ld,
                None
                if (slope_delta_area_ll is None or slope_delta_area_ld is None)
                else slope_delta_area_ld - slope_delta_area_ll,
                slope_delta_ecc_ll,
                slope_delta_ecc_ld,
                None
                if (slope_delta_ecc_ll is None or slope_delta_ecc_ld is None)
                else slope_delta_ecc_ld - slope_delta_ecc_ll,
            ]
        )

    return output


def write_sheet(ws, header, rows):
    ws.append(header)
    for row in rows:
        ws.append(row)


def main():
    if not CSV_LL.exists() or not CSV_LD.exists():
        raise FileNotFoundError(f"Expected input files: {CSV_LL} and {CSV_LD}")

    rows_ll = load_rows(CSV_LL)
    rows_ld = load_rows(CSV_LD)

    wb = Workbook()

    ws_conclusive = wb.active
    ws_conclusive.title = "conclusive"
    write_sheet(
        ws_conclusive,
        [
            "parameter",
            "fold_change",
            "param_value_ll",
            "param_value_ld",
            "base_param_value_ll",
            "base_param_value_ld",
            "mean_area_ll",
            "mean_area_ld",
            "mean_eccentricity_ll",
            "mean_eccentricity_ld",
            "mean_delta_area_ll",
            "mean_delta_area_ld",
            "mean_delta_eccentricity_ll",
            "mean_delta_eccentricity_ld",
            "delta_mean_area_ld_minus_ll",
            "delta_mean_eccentricity_ld_minus_ll",
            "delta_mean_delta_area_ld_minus_ll",
            "delta_mean_delta_eccentricity_ld_minus_ll",
        ],
        build_conclusive_rows(rows_ll, rows_ld),
    )

    ws_ll = wb.create_sheet("ll")
    write_sheet(
        ws_ll,
        [
            "parameter",
            "param_value",
            "base_param_value",
            "fold_change",
            "mean_area",
            "mean_eccentricity",
            "mean_delta_area",
            "mean_delta_eccentricity",
            "condition",
        ],
        prepare_condition_sheet_rows(rows_ll, "ll"),
    )

    ws_ld = wb.create_sheet("ld")
    write_sheet(
        ws_ld,
        [
            "parameter",
            "param_value",
            "base_param_value",
            "fold_change",
            "mean_area",
            "mean_eccentricity",
            "mean_delta_area",
            "mean_delta_eccentricity",
            "condition",
        ],
        prepare_condition_sheet_rows(rows_ld, "ld"),
    )

    ws_rate = wb.create_sheet("rate_summary")
    write_sheet(
        ws_rate,
        [
            "parameter",
            "slope_mean_delta_area_vs_fold_change_ll",
            "slope_mean_delta_area_vs_fold_change_ld",
            "delta_slope_mean_delta_area_ld_minus_ll",
            "slope_mean_delta_eccentricity_vs_fold_change_ll",
            "slope_mean_delta_eccentricity_vs_fold_change_ld",
            "delta_slope_mean_delta_eccentricity_ld_minus_ll",
        ],
        build_rate_summary_rows(rows_ll, rows_ld),
    )

    wb.save(OUT_XLSX)
    print(f"Saved Excel file: {OUT_XLSX}")


if __name__ == "__main__":
    main()
